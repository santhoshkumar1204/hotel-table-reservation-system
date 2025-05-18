from openpyxl import load_workbook


def get_and_edit_value_in_excel(file_path, sheet_name, date, slot):
    try:
        # Load the workbook and the specified sheet
        workbook = load_workbook(filename=file_path)
        if sheet_name not in workbook.sheetnames:
            print('invalid sheet name')

        sheet = workbook[sheet_name]

        # Find the row and column for the given date and slot
        date_row = None
        slot_column = None

        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == slot:
                    slot_column = cell.column

        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == date:
                    date_row = cell.row

        # Check if both date and slot were found
        if slot_column is None or date_row is None:
            status = "invalid input"
            return status

        # Access the original value
        original_value = sheet.cell(row=date_row, column=slot_column).value
        if original_value > 0:
            sheet.cell(row=date_row, column=slot_column).value -= 1
            status = "available"

        # Update the cell with the new value
        else:
            status = "not available"

        # Save the workbook
        workbook.save(file_path)

    except FileNotFoundError:
        status = "invalid input"

    return status
# Example usage


def get_and_minus_value_in_excel(file_path, sheet_name, date, slot):
    try:
        # Load the workbook and the specified sheet
        workbook = load_workbook(filename=file_path)
        if sheet_name not in workbook.sheetnames:
            print('invalid sheet name')

        sheet = workbook[sheet_name]

        # Find the row and column for the given date and slot
        date_row = None
        slot_column = None

        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == slot:
                    slot_column = cell.column

        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == date:
                    date_row = cell.row

        # Check if both date and slot were found
        if slot_column is None or date_row is None:
            status = "invalid input"
            return status

        # Access the original value
        original_value = sheet.cell(row=date_row, column=slot_column).value
        if original_value < 5:
            sheet.cell(row=date_row, column=slot_column).value += 1
            status = "success"

        # Update the cell with the new value
        else:
            status = "wrong"

        # Save the workbook
        workbook.save(file_path)

    except FileNotFoundError:
        status = "invalid input"

    return status