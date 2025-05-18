import openpyxl
from datetime import datetime, timedelta


def update_schedule(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Calculate yesterday's date and the tenth day from today
        today = datetime.today().date()
        yesterday = today - timedelta(days=1)
        tenth_day = today + timedelta(days=9)

        # Convert dates to strings in the format "dd/mm/yyyy"
        yesterday_str = yesterday.strftime("%d/%m/%Y")
        tenth_day_str = tenth_day.strftime("%d/%m/%Y")

        # Find and delete the row with yesterday's date
        yesterday_row = None
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and cell_value == yesterday_str:
                yesterday_row = row
                break

        if yesterday_row:
            ws.delete_rows(yesterday_row)
        # Check if the tenth day already exists
        tenth_day_exists = False
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and cell_value == tenth_day_str:
                tenth_day_exists = True
                break

        # Insert a new row at the end with the tenth day from today if it doesn't exist
        if not tenth_day_exists:
            new_row = ws.max_row + 1
            ws.insert_rows(new_row)

            # Set the new date and fill other columns with default value 5
            ws.cell(row=new_row, column=1).value = tenth_day_str
            for col in range(2, ws.max_column + 1):
                ws.cell(row=new_row, column=col).value = 5

    # Save the workbook
    wb.save(file_path)


# Usage
filepath = 'schedule.xlsx'
update_schedule(filepath)
